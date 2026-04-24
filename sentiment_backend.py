from __future__ import annotations

import io
import re
from dataclasses import dataclass
from typing import Dict, Iterable, List, Optional, Tuple, Any

import pandas as pd


RESULTS_SHEET_NAME = "Results"
CITATIONS_SHEET_NAME = "Citations"

PROMPT_COL_SEGMENTATION = "Prompt"
AREA_COL_SEGMENTATION = "Sentiment area"
QUESTION_COL_RESULTS = "question"
SENTIMENT_COL_RESULTS = "sentiment"
URL_COL_CITATIONS = "url"
DOMAIN_COL_CITATIONS = "domain"
TITLE_COL_CITATIONS = "title"

PROMPT_LEVEL_SHEET = "Prompt Level Averages"
AREA_SUMMARY_SHEET = "Sentiment Area Summary"
TOP_DOMAINS_TOTAL_SHEET = "Top Domains Total"
TOP_URLS_TOTAL_SHEET = "Top URLs Total"
TOP_DOMAINS_BY_AREA_SHEET = "Top Domains By Area"
TOP_URLS_BY_AREA_SHEET = "Top URLs By Area"
MATCHED_ROWS_SHEET = "Matched Results"
MATCHED_CITATIONS_SHEET = "Matched Citations"
UNMATCHED_ROWS_SHEET = "Unmatched Results"
UNMATCHED_CITATIONS_SHEET = "Unmatched Citations"

PLACEHOLDER_PATTERN = re.compile(
    r"\[(?:brand|varum[aä]rke|varumaerke|varumarke|mærke|merke|marca|marque)\]",
    flags=re.IGNORECASE,
)


@dataclass(frozen=True)
class PromptPattern:
    original_prompt: str
    output_prompt: str
    sentiment_area: str
    regex: re.Pattern
    sort_order: int


class SentimentKingError(Exception):
    """Raised for user-facing validation errors."""


def _reset_file(file_obj) -> None:
    try:
        file_obj.seek(0)
    except Exception:
        pass


def _canonical_col_map(columns: Iterable[object]) -> Dict[str, object]:
    return {str(c).strip().lower(): c for c in columns}


def _require_column(df: pd.DataFrame, wanted: str, source_name: str) -> object:
    col_map = _canonical_col_map(df.columns)
    key = wanted.strip().lower()
    if key not in col_map:
        available = ", ".join(map(str, df.columns))
        raise SentimentKingError(
            f"Missing required column '{wanted}' in {source_name}. Available columns: {available}"
        )
    return col_map[key]


def _optional_column(df: pd.DataFrame, wanted: str) -> Optional[object]:
    return _canonical_col_map(df.columns).get(wanted.strip().lower())


def _normalize_text_for_matching(value: object) -> str:
    if pd.isna(value):
        return ""
    text = str(value)
    text = text.replace("’", "'").replace("‘", "'").replace("`", "'")
    text = text.replace("\u00a0", " ")
    text = re.sub(r"\s+", " ", text).strip()
    return text


def _placeholder_to_output_prompt(prompt: str) -> str:
    return PLACEHOLDER_PATTERN.sub("Brand", prompt)


def _compile_prompt_regex(prompt: str) -> re.Pattern:
    """Convert a segmented prompt with [Brand]/[Varumärke] into a regex.

    The regex captures the brand value from uploaded Results/Citations questions.
    Whitespace is flexible; punctuation/text must otherwise match.
    """
    normalized = _normalize_text_for_matching(prompt)
    matches = list(PLACEHOLDER_PATTERN.finditer(normalized))
    if not matches:
        raise SentimentKingError(
            f"Prompt is missing a brand placeholder like [Brand] or [Varumärke]: {prompt}"
        )

    parts: List[str] = []
    last = 0
    for idx, match in enumerate(matches):
        literal = normalized[last : match.start()]
        parts.append(_escape_literal_with_flexible_spaces(literal))
        if idx == 0:
            parts.append(r"(?P<brand>.+?)")
        else:
            parts.append(r".+?")
        last = match.end()
    parts.append(_escape_literal_with_flexible_spaces(normalized[last:]))

    pattern = "^" + "".join(parts) + "$"
    return re.compile(pattern, flags=re.IGNORECASE)


def _escape_literal_with_flexible_spaces(text: str) -> str:
    escaped = re.escape(text)
    escaped = escaped.replace(r"\ ", r"\s+").replace(" ", r"\s+")
    return escaped


def _clean_brand(raw_brand: object) -> str:
    brand = _normalize_text_for_matching(raw_brand)
    brand = brand.strip(" \t\n\r'\".,;:!?()[]{}")
    brand = re.sub(r"\s+", " ", brand)
    return brand


def read_segmentation_file(file_obj) -> pd.DataFrame:
    _reset_file(file_obj)
    try:
        df = pd.read_excel(file_obj)
    except Exception as exc:
        raise SentimentKingError(f"Could not read segmentation Excel file: {exc}") from exc

    prompt_col = _require_column(df, PROMPT_COL_SEGMENTATION, "segmentation file")
    area_col = _require_column(df, AREA_COL_SEGMENTATION, "segmentation file")

    df = df[[prompt_col, area_col]].copy()
    df.columns = [PROMPT_COL_SEGMENTATION, AREA_COL_SEGMENTATION]
    df[PROMPT_COL_SEGMENTATION] = df[PROMPT_COL_SEGMENTATION].map(_normalize_text_for_matching)
    df[AREA_COL_SEGMENTATION] = df[AREA_COL_SEGMENTATION].astype(str).str.strip()
    df = df[(df[PROMPT_COL_SEGMENTATION] != "") & (df[AREA_COL_SEGMENTATION] != "")]

    if df.empty:
        raise SentimentKingError("Segmentation file has no usable Prompt/Sentiment area rows.")

    return df.reset_index(drop=True)


def build_prompt_patterns(segmentation_df: pd.DataFrame) -> List[PromptPattern]:
    patterns: List[PromptPattern] = []
    seen: set[Tuple[str, str]] = set()

    for idx, row in segmentation_df.iterrows():
        original_prompt = _normalize_text_for_matching(row[PROMPT_COL_SEGMENTATION])
        sentiment_area = str(row[AREA_COL_SEGMENTATION]).strip()
        output_prompt = _placeholder_to_output_prompt(original_prompt)
        key = (output_prompt.lower(), sentiment_area.lower())
        if key in seen:
            continue
        seen.add(key)
        patterns.append(
            PromptPattern(
                original_prompt=original_prompt,
                output_prompt=output_prompt,
                sentiment_area=sentiment_area,
                regex=_compile_prompt_regex(original_prompt),
                sort_order=idx,
            )
        )

    if not patterns:
        raise SentimentKingError("No valid prompt patterns could be built from the segmentation file.")

    return sorted(patterns, key=lambda p: (-len(p.original_prompt), p.sort_order))


def read_results_files(files: Iterable) -> pd.DataFrame:
    frames: List[pd.DataFrame] = []
    for file_obj in files:
        _reset_file(file_obj)
        file_name = getattr(file_obj, "name", "uploaded file")
        try:
            df = pd.read_excel(file_obj, sheet_name=RESULTS_SHEET_NAME)
        except ValueError as exc:
            raise SentimentKingError(f"'{file_name}' does not contain a '{RESULTS_SHEET_NAME}' sheet.") from exc
        except Exception as exc:
            raise SentimentKingError(f"Could not read '{file_name}': {exc}") from exc

        question_col = _require_column(df, QUESTION_COL_RESULTS, f"{file_name} / {RESULTS_SHEET_NAME}")
        sentiment_col = _require_column(df, SENTIMENT_COL_RESULTS, f"{file_name} / {RESULTS_SHEET_NAME}")

        keep_cols = [question_col, sentiment_col]
        optional_cols = []
        for optional in ["platform", "model", "run_date", "country"]:
            optional_col = _optional_column(df, optional)
            if optional_col is not None:
                optional_cols.append(optional_col)
        keep_cols += optional_cols

        small = df[keep_cols].copy()
        rename = {question_col: QUESTION_COL_RESULTS, sentiment_col: SENTIMENT_COL_RESULTS}
        for col in optional_cols:
            rename[col] = str(col).strip().lower()
        small = small.rename(columns=rename)
        small["source_file"] = file_name
        frames.append(small)

    if not frames:
        raise SentimentKingError("Upload at least one results Excel file.")

    combined = pd.concat(frames, ignore_index=True)
    combined[QUESTION_COL_RESULTS] = combined[QUESTION_COL_RESULTS].map(_normalize_text_for_matching)
    combined[SENTIMENT_COL_RESULTS] = pd.to_numeric(combined[SENTIMENT_COL_RESULTS], errors="coerce")
    combined = combined[(combined[QUESTION_COL_RESULTS] != "") & combined[SENTIMENT_COL_RESULTS].notna()]

    if combined.empty:
        raise SentimentKingError("No usable question/sentiment rows found in the uploaded results files.")

    return combined.reset_index(drop=True)


def read_citation_files(files: Iterable) -> pd.DataFrame:
    frames: List[pd.DataFrame] = []
    missing_citations: List[str] = []

    for file_obj in files:
        _reset_file(file_obj)
        file_name = getattr(file_obj, "name", "uploaded file")
        try:
            df = pd.read_excel(file_obj, sheet_name=CITATIONS_SHEET_NAME)
        except ValueError:
            missing_citations.append(file_name)
            continue
        except Exception as exc:
            raise SentimentKingError(f"Could not read '{file_name}' / '{CITATIONS_SHEET_NAME}': {exc}") from exc

        question_col = _require_column(df, QUESTION_COL_RESULTS, f"{file_name} / {CITATIONS_SHEET_NAME}")
        url_col = _require_column(df, URL_COL_CITATIONS, f"{file_name} / {CITATIONS_SHEET_NAME}")
        domain_col = _require_column(df, DOMAIN_COL_CITATIONS, f"{file_name} / {CITATIONS_SHEET_NAME}")

        keep_cols = [question_col, url_col, domain_col]
        rename = {
            question_col: QUESTION_COL_RESULTS,
            url_col: URL_COL_CITATIONS,
            domain_col: DOMAIN_COL_CITATIONS,
        }

        for optional in [TITLE_COL_CITATIONS, "platform", "model", "run_date", "country", "source"]:
            optional_col = _optional_column(df, optional)
            if optional_col is not None:
                keep_cols.append(optional_col)
                rename[optional_col] = str(optional_col).strip().lower()

        small = df[keep_cols].copy().rename(columns=rename)
        small["source_file"] = file_name
        frames.append(small)

    if not frames:
        empty = pd.DataFrame(columns=[QUESTION_COL_RESULTS, URL_COL_CITATIONS, DOMAIN_COL_CITATIONS, "source_file"])
        empty.attrs["missing_citations_files"] = missing_citations
        return empty

    combined = pd.concat(frames, ignore_index=True)
    combined[QUESTION_COL_RESULTS] = combined[QUESTION_COL_RESULTS].map(_normalize_text_for_matching)
    combined[URL_COL_CITATIONS] = combined[URL_COL_CITATIONS].map(_normalize_text_for_matching)
    combined[DOMAIN_COL_CITATIONS] = combined[DOMAIN_COL_CITATIONS].map(_normalize_text_for_matching).str.lower()
    combined = combined[(combined[QUESTION_COL_RESULTS] != "") & ((combined[URL_COL_CITATIONS] != "") | (combined[DOMAIN_COL_CITATIONS] != ""))]
    combined.attrs["missing_citations_files"] = missing_citations
    return combined.reset_index(drop=True)


def _match_question(question: object, patterns: List[PromptPattern]) -> Optional[dict]:
    normalized_question = _normalize_text_for_matching(question)
    for pattern in patterns:
        match = pattern.regex.match(normalized_question)
        if match:
            brand = _clean_brand(match.groupdict().get("brand", ""))
            if not brand:
                continue
            return {
                "Prompt": pattern.output_prompt,
                "Sentiment area": pattern.sentiment_area,
                "Brand": brand,
                "question": normalized_question,
            }
    return None


def match_results_to_segments(results_df: pd.DataFrame, patterns: List[PromptPattern]) -> Tuple[pd.DataFrame, pd.DataFrame]:
    matched_rows: List[dict] = []
    unmatched_rows: List[dict] = []

    for _, row in results_df.iterrows():
        match_data = _match_question(row[QUESTION_COL_RESULTS], patterns)
        if match_data is None:
            unmatched_rows.append(row.to_dict())
            continue

        match_data.update(
            {
                "sentiment": row[SENTIMENT_COL_RESULTS],
                "source_file": row.get("source_file", ""),
            }
        )
        for optional in ["platform", "model", "run_date", "country"]:
            if optional in row.index:
                match_data[optional] = row.get(optional, "")
        matched_rows.append(match_data)

    matched_df = pd.DataFrame(matched_rows)
    unmatched_df = pd.DataFrame(unmatched_rows)

    if matched_df.empty:
        raise SentimentKingError(
            "None of the Results questions matched the segmentation prompts. Check that the wording is identical except for the brand placeholder."
        )

    return matched_df, unmatched_df


def match_citations_to_segments(citations_df: pd.DataFrame, patterns: List[PromptPattern]) -> Tuple[pd.DataFrame, pd.DataFrame]:
    if citations_df.empty:
        return pd.DataFrame(), citations_df.copy()

    matched_rows: List[dict] = []
    unmatched_rows: List[dict] = []

    for _, row in citations_df.iterrows():
        match_data = _match_question(row[QUESTION_COL_RESULTS], patterns)
        if match_data is None:
            unmatched_rows.append(row.to_dict())
            continue

        match_data.update(
            {
                "url": row.get(URL_COL_CITATIONS, ""),
                "domain": row.get(DOMAIN_COL_CITATIONS, ""),
                "source_file": row.get("source_file", ""),
            }
        )
        for optional in [TITLE_COL_CITATIONS, "platform", "model", "run_date", "country", "source"]:
            if optional in row.index:
                match_data[optional] = row.get(optional, "")
        matched_rows.append(match_data)

    return pd.DataFrame(matched_rows), pd.DataFrame(unmatched_rows)


def make_prompt_level_averages(matched_df: pd.DataFrame, segmentation_df: pd.DataFrame) -> pd.DataFrame:
    grouped = (
        matched_df.groupby(["Prompt", "Sentiment area", "Brand"], dropna=False)["sentiment"]
        .mean()
        .reset_index()
    )

    pivot = grouped.pivot_table(
        index=["Prompt", "Sentiment area"],
        columns="Brand",
        values="sentiment",
        aggfunc="mean",
    ).reset_index()

    brand_cols = sorted([c for c in pivot.columns if c not in ["Prompt", "Sentiment area"]], key=str.casefold)
    pivot = pivot[["Prompt", "Sentiment area"] + brand_cols]

    order_df = segmentation_df.copy()
    order_df["Prompt"] = order_df[PROMPT_COL_SEGMENTATION].map(_placeholder_to_output_prompt)
    order_df = order_df.rename(columns={AREA_COL_SEGMENTATION: "Sentiment area"})
    order_df["__order"] = range(len(order_df))
    order_df = order_df[["Prompt", "Sentiment area", "__order"]].drop_duplicates()

    pivot = pivot.merge(order_df, on=["Prompt", "Sentiment area"], how="left")
    pivot = pivot.sort_values(["__order", "Sentiment area", "Prompt"], na_position="last").drop(columns="__order")
    return pivot.reset_index(drop=True)


def make_area_summary(prompt_level_df: pd.DataFrame) -> pd.DataFrame:
    brand_cols = [c for c in prompt_level_df.columns if c not in ["Prompt", "Sentiment area"]]
    if not brand_cols:
        raise SentimentKingError("No brand columns were produced. Check prompt matching and brand extraction.")

    summary = prompt_level_df.groupby("Sentiment area", sort=False)[brand_cols].mean().reset_index()
    summary = summary.rename(columns={"Sentiment area": "Sentiment Area"})
    summary["Industry Average"] = summary[brand_cols].mean(axis=1, skipna=True)
    return summary


def _rank_counts(
    df: pd.DataFrame,
    value_col: str,
    group_cols: List[str],
    top_n: int,
    extra_cols: Optional[List[str]] = None,
) -> pd.DataFrame:
    if df.empty or value_col not in df.columns:
        return pd.DataFrame()

    filtered = df[df[value_col].map(_normalize_text_for_matching) != ""].copy()
    if filtered.empty:
        return pd.DataFrame()

    grouped = filtered.groupby(group_cols + [value_col], dropna=False).size().reset_index(name="Citations")
    grouped = grouped.sort_values(group_cols + ["Citations", value_col], ascending=[True] * len(group_cols) + [False, True])
    grouped["Rank"] = grouped.groupby(group_cols)["Citations"].rank(method="first", ascending=False).astype(int)
    grouped = grouped[grouped["Rank"] <= top_n]

    if extra_cols:
        for extra_col in extra_cols:
            if extra_col in filtered.columns and extra_col not in grouped.columns:
                first_values = (
                    filtered[[*group_cols, value_col, extra_col]]
                    .dropna(subset=[extra_col])
                    .drop_duplicates(subset=[*group_cols, value_col])
                )
                grouped = grouped.merge(first_values, on=[*group_cols, value_col], how="left")

    ordered_cols = [*group_cols, "Rank", value_col, "Citations"]
    if extra_cols:
        ordered_cols += [c for c in extra_cols if c in grouped.columns]
    return grouped[ordered_cols].reset_index(drop=True)


def make_citation_summaries(matched_citations_df: pd.DataFrame) -> Dict[str, pd.DataFrame]:
    if matched_citations_df.empty:
        empty = pd.DataFrame()
        return {
            "top_domains_total": empty,
            "top_urls_total": empty,
            "top_domains_by_area": empty,
            "top_urls_by_area": empty,
        }

    top_domains_total = _rank_counts(
        matched_citations_df,
        value_col="domain",
        group_cols=["Brand"],
        top_n=3,
    ).rename(columns={"domain": "Domain"})

    top_urls_total = _rank_counts(
        matched_citations_df,
        value_col="url",
        group_cols=["Brand"],
        top_n=10,
        extra_cols=["domain", "title"],
    ).rename(columns={"url": "URL", "domain": "Domain", "title": "Title"})

    top_domains_by_area = _rank_counts(
        matched_citations_df,
        value_col="domain",
        group_cols=["Sentiment area", "Brand"],
        top_n=3,
    ).rename(columns={"Sentiment area": "Sentiment Area", "domain": "Domain"})

    top_urls_by_area = _rank_counts(
        matched_citations_df,
        value_col="url",
        group_cols=["Sentiment area", "Brand"],
        top_n=10,
        extra_cols=["domain", "title"],
    ).rename(columns={"Sentiment area": "Sentiment Area", "url": "URL", "domain": "Domain", "title": "Title"})

    return {
        "top_domains_total": top_domains_total,
        "top_urls_total": top_urls_total,
        "top_domains_by_area": top_domains_by_area,
        "top_urls_by_area": top_urls_by_area,
    }


def build_output_workbook(
    segmentation_file,
    results_files: Iterable,
    decimals: Optional[int] = None,
) -> Tuple[bytes, Dict[str, Any]]:
    results_files = list(results_files)
    segmentation_df = read_segmentation_file(segmentation_file)
    patterns = build_prompt_patterns(segmentation_df)
    results_df = read_results_files(results_files)
    matched_df, unmatched_df = match_results_to_segments(results_df, patterns)

    citations_df = read_citation_files(results_files)
    matched_citations_df, unmatched_citations_df = match_citations_to_segments(citations_df, patterns)
    citation_summaries = make_citation_summaries(matched_citations_df)

    prompt_level = make_prompt_level_averages(matched_df, segmentation_df)
    area_summary = make_area_summary(prompt_level)

    if decimals is not None:
        numeric_cols_prompt = prompt_level.select_dtypes(include="number").columns
        numeric_cols_summary = area_summary.select_dtypes(include="number").columns
        prompt_level[numeric_cols_prompt] = prompt_level[numeric_cols_prompt].round(decimals)
        area_summary[numeric_cols_summary] = area_summary[numeric_cols_summary].round(decimals)

    output = io.BytesIO()
    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        prompt_level.to_excel(writer, sheet_name=PROMPT_LEVEL_SHEET, index=False)
        area_summary.to_excel(writer, sheet_name=AREA_SUMMARY_SHEET, index=False)
        citation_summaries["top_domains_total"].to_excel(writer, sheet_name=TOP_DOMAINS_TOTAL_SHEET, index=False)
        citation_summaries["top_urls_total"].to_excel(writer, sheet_name=TOP_URLS_TOTAL_SHEET, index=False)
        citation_summaries["top_domains_by_area"].to_excel(writer, sheet_name=TOP_DOMAINS_BY_AREA_SHEET, index=False)
        citation_summaries["top_urls_by_area"].to_excel(writer, sheet_name=TOP_URLS_BY_AREA_SHEET, index=False)
        matched_df.to_excel(writer, sheet_name=MATCHED_ROWS_SHEET, index=False)
        if not matched_citations_df.empty:
            matched_citations_df.to_excel(writer, sheet_name=MATCHED_CITATIONS_SHEET, index=False)
        if not unmatched_df.empty:
            unmatched_df.to_excel(writer, sheet_name=UNMATCHED_ROWS_SHEET, index=False)
        if not unmatched_citations_df.empty:
            unmatched_citations_df.to_excel(writer, sheet_name=UNMATCHED_CITATIONS_SHEET, index=False)

        _format_workbook(writer.book)

    missing_citation_files = citations_df.attrs.get("missing_citations_files", []) if hasattr(citations_df, "attrs") else []
    stats = {
        "results_rows_used": int(len(results_df)),
        "matched_rows": int(len(matched_df)),
        "unmatched_rows": int(len(unmatched_df)),
        "brands": int(len([c for c in prompt_level.columns if c not in ["Prompt", "Sentiment area"]])),
        "sentiment_areas": int(area_summary["Sentiment Area"].nunique()),
        "citation_rows_used": int(len(citations_df)),
        "matched_citation_rows": int(len(matched_citations_df)),
        "unmatched_citation_rows": int(len(unmatched_citations_df)),
        "missing_citation_files": missing_citation_files,
        "prompt_level_preview": prompt_level,
        "area_summary_preview": area_summary,
        "top_domains_total_preview": citation_summaries["top_domains_total"],
        "top_urls_total_preview": citation_summaries["top_urls_total"],
        "top_domains_by_area_preview": citation_summaries["top_domains_by_area"],
        "top_urls_by_area_preview": citation_summaries["top_urls_by_area"],
    }
    return output.getvalue(), stats


def _format_workbook(workbook) -> None:
    from openpyxl.styles import Font, PatternFill
    from openpyxl.utils import get_column_letter

    header_fill = PatternFill("solid", fgColor="D9EAF7")
    header_font = Font(bold=True)

    for ws in workbook.worksheets:
        ws.freeze_panes = "A2"
        ws.auto_filter.ref = ws.dimensions
        for cell in ws[1]:
            cell.font = header_font
            cell.fill = header_fill

        for column_cells in ws.columns:
            max_length = 0
            col_idx = column_cells[0].column
            col_letter = get_column_letter(col_idx)
            for cell in column_cells[:200]:
                if cell.value is None:
                    continue
                max_length = max(max_length, len(str(cell.value)))
            ws.column_dimensions[col_letter].width = min(max(max_length + 2, 12), 70)

        for row in ws.iter_rows(min_row=2):
            for cell in row:
                if isinstance(cell.value, (int, float)):
                    cell.number_format = "0.000"
